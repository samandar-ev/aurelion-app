import json
from decimal import Decimal, InvalidOperation
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.views import LoginView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import TemplateView, ListView, DetailView
from django.urls import reverse_lazy
from django.views.generic import TemplateView, ListView, DetailView, UpdateView
from django.urls import reverse_lazy
from django.views import View
from django.contrib import messages
from django.db import transaction
from django.http import JsonResponse, HttpResponseBadRequest
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.utils import timezone
from django.db.models import Sum, F, Q, ProtectedError
from PIL import Image
import random
import string
from .models import Product, Client, ProductImage, Barcode, ProductVariant, Location, Order, OrderItem, Return, ReturnItem, Promotion, PromotionUsage
from .forms import ProductForm, ClientForm, LuxuryProductForm, ProductVariantFormSet, BRAND_CHOICES, MATERIAL_CHOICES, BRAND_SUGGESTIONS, COLOR_SUGGESTIONS
from .permissions import (
    OwnerRequiredMixin, CashierRequiredMixin, SalesAssociateRequiredMixin
)
from django.views.generic.edit import CreateView

class CustomLoginView(LoginView):
    template_name = 'core/login.html'
    redirect_authenticated_user = True
    
    def get_success_url(self):
        # redirect based on the role of user
        user = self.request.user
        if user.is_sales_associate() and not user.is_cashier():
            return reverse_lazy('associate_dashboard')
        return reverse_lazy('dashboard')

class DashboardView(LoginRequiredMixin, TemplateView):
    template_name = 'core/dashboard.html'
    
    def get_context_data(self, **kwargs):
        # this is quite big but shows stats on main page
        context = super().get_context_data(**kwargs)
        today = timezone.localdate()
        
        gross_sales = Decimal('0.00')
        refund_amount = Decimal('0.00')
        net_sales = Decimal('0.00')
        items_sold = 0
        today_count = 0
        recent_orders = []
        top_products = []
        low_stock = []
        low_stock_count = 0
        active_promotions = []
        db_error = None
        
        try:
            today_orders = Order.objects.filter(
                created_at__date=today, 
                status=Order.Status.COMPLETED,
                type=Order.Type.SALE
            )
            
            today_returns = Return.objects.filter(created_at__date=today)
            for ret in today_returns:
                for item in ret.items.all():
                    refund_amount += item.order_item.unit_price * item.quantity
            
            gross_sales = today_orders.aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')
            net_sales = gross_sales - refund_amount
            
            items_sold = OrderItem.objects.filter(order__in=today_orders).aggregate(qty=Sum('quantity'))['qty'] or 0
            
            orders_query = Order.objects.all()
            search_query = self.request.GET.get('search', '').strip()
            if search_query:
                orders_query = orders_query.filter(
                    Q(order_code__icontains=search_query) | Q(id__icontains=search_query)
                )
            recent_orders = list(orders_query.order_by('-created_at')[:50])
            
            top_products = list(
                OrderItem.objects
                .filter(order__in=today_orders)
                .select_related('variant__product')
                .values('variant__product__id', 'variant__product__name', 'variant__product__brand')
                .annotate(total_qty=Sum('quantity'), total_sales=Sum(F('quantity') * F('unit_price')))
                .order_by('-total_qty')[:5]
            )
            
            low_stock = list(
                ProductVariant.objects
                .filter(initial_quantity__lt=5, product__is_archived=False)
                .select_related('product')
                .order_by('-updated_at')[:5]
            )
            
            low_stock_count = ProductVariant.objects.filter(
                initial_quantity__lt=5, 
                product__is_archived=False
            ).count()
            
            active_promotions = list(
                Promotion.objects.filter(
                    is_active=True,
                    start_date__lte=timezone.now(),
                    end_date__gte=timezone.now()
                ).order_by('-created_at')[:3]
            )
            
            today_count = today_orders.count()
            
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Database error in dashboard: {str(e)}", exc_info=True)
            db_error = str(e)
            messages.error(self.request, f"Unable to load some dashboard data: {str(e)}")
        
        context.update({
            'gross_sales': gross_sales,
            'refund_amount': refund_amount,
            'net_sales': net_sales,
            'items_sold': items_sold,
            'today_count': today_count,
            'recent_orders': recent_orders,
            'top_products': top_products,
            'low_stock': low_stock,
            'low_stock_count': low_stock_count,
            'active_promotions': active_promotions,
            'db_error': db_error,
        })
        return context

class OrderListView(CashierRequiredMixin, ListView):
    model = Order
    template_name = 'core/order_list.html'
    context_object_name = 'orders'
    paginate_by = 50
    
    def get_queryset(self):
        return Order.objects.order_by('-created_at')

class OrderDetailView(LoginRequiredMixin, DetailView):
    model = Order
    template_name = 'core/order_detail.html'
    context_object_name = 'order'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        order = self.get_object()
        items = order.items.select_related('variant__product').all()
        context['items'] = items
        
        import json
        order_data = {
            'order_code': order.order_code or str(order.id),
            'client': str(order.client) if order.client else 'Walk-in',
            'cashier': order.created_by.username,
            'created_at': order.created_at.isoformat(),
            'total': float(order.total_amount),
            'discount': float(order.total_discount),
            'subtotal': float(order.total_amount + order.total_discount),
            'items': [
                {
                    'name': f"{item.variant.product.brand} {item.variant.product.name}",
                    'sku': item.variant.sku,
                    'color': item.variant.color or '',
                    'size': item.variant.size or '',
                    'qty': item.quantity,
                    'price': float(item.unit_price),
                    'total': float(item.line_total)
                }
                for item in items
            ]
        }
        context['order_data_json'] = json.dumps(order_data)
        return context

class ProductListView(SalesAssociateRequiredMixin, ListView):
    model = Product
    template_name = 'core/product_list.html'
    context_object_name = 'products'
    paginate_by = 20

    def get_queryset(self):
        queryset = super().get_queryset().filter(is_archived=False)
        
        q = self.request.GET.get('q')
        if q:
            variant_products = Product.objects.filter(variants__sku__icontains=q)
            queryset = (queryset.filter(name__icontains=q) | queryset.filter(brand__icontains=q) | queryset.filter(base_sku__icontains=q) | variant_products).distinct()
        
        category = self.request.GET.get('category')
        if category:
            queryset = queryset.filter(category=category)
        
        brand = self.request.GET.get('brand')
        if brand:
            queryset = queryset.filter(brand=brand)
        
        season = self.request.GET.get('season')
        if season:
            queryset = queryset.filter(season=season)
        
        material = self.request.GET.get('material')
        if material:
            queryset = queryset.filter(material=material)
        
        sort = self.request.GET.get('sort', '')
        if sort == 'name_asc':
            queryset = queryset.order_by('name')
        elif sort == 'name_desc':
            queryset = queryset.order_by('-name')
        elif sort == 'price_asc':
            queryset = queryset.annotate(min_price=Sum('variants__retail_price')).order_by('min_price')
        elif sort == 'price_desc':
            queryset = queryset.annotate(min_price=Sum('variants__retail_price')).order_by('-min_price')
        elif sort == 'date_desc':
            queryset = queryset.order_by('-id')                                     
        elif sort == 'date_asc':
            queryset = queryset.order_by('id')
        elif sort == 'stock_asc':
            queryset = queryset.annotate(total_stock=Sum('variants__initial_quantity')).order_by('total_stock')
        elif sort == 'stock_desc':
            queryset = queryset.annotate(total_stock=Sum('variants__initial_quantity')).order_by('-total_stock')
        else:
            queryset = queryset.order_by('-id')                         
        
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        q = self.request.GET.get('q')
        if q:
            matching_variants = ProductVariant.objects.select_related('product').filter(sku__icontains=q)
            context['variant_matches'] = matching_variants
            context['search_term'] = q
        
        context['brands'] = Product.objects.filter(is_archived=False).exclude(brand='').values_list('brand', flat=True).distinct().order_by('brand')
        context['seasons'] = Product.objects.filter(is_archived=False).exclude(season='').values_list('season', flat=True).distinct().order_by('season')
        context['materials'] = Product.objects.filter(is_archived=False).exclude(material='').values_list('material', flat=True).distinct().order_by('material')
        
        active_filters = sum([
            bool(self.request.GET.get('category')),
            bool(self.request.GET.get('brand')),
            bool(self.request.GET.get('season')),
            bool(self.request.GET.get('material')),
        ])
        context['active_filters'] = active_filters if active_filters > 0 else None
        
        return context


class ArchivedProductListView(LoginRequiredMixin, ListView):
    model = Product
    template_name = 'core/product_list.html'
    context_object_name = 'products'
    paginate_by = 20

    def get_queryset(self):
        queryset = Product.objects.filter(is_archived=True)
        q = self.request.GET.get('q')
        if q:
            variant_products = Product.objects.filter(variants__sku__icontains=q)
            queryset = (queryset.filter(name__icontains=q) | queryset.filter(brand__icontains=q) | queryset.filter(base_sku__icontains=q) | variant_products).distinct()
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['is_archived_view'] = True
        return context

class ArchiveProductView(CashierRequiredMixin, View):
    def post(self, request, pk):
        product = get_object_or_404(Product, pk=pk)
        product.is_archived = True
        product.save()
        messages.success(request, f"Product {product.name} archived.")
        return redirect('product_list')

class UnarchiveProductView(OwnerRequiredMixin, View):
    def post(self, request, pk):
        product = get_object_or_404(Product, pk=pk)
        product.is_archived = False
        product.save()
        messages.success(request, f"Product {product.name} restored from archive.")
        return redirect('archived_product_list')


class BarcodeGeneratorView(CashierRequiredMixin, ListView):
    """View to select products and generate barcode PDF"""
    model = Product
    template_name = 'core/barcode_generator.html'
    context_object_name = 'products'
    paginate_by = 50

    def get_queryset(self):
        queryset = super().get_queryset().filter(is_archived=False).prefetch_related('variants')
        q = self.request.GET.get('q')
        if q:
            variant_products = Product.objects.filter(variants__sku__icontains=q)
            queryset = (queryset.filter(name__icontains=q) | queryset.filter(brand__icontains=q) | queryset.filter(base_sku__icontains=q) | variant_products).distinct()
        return queryset.order_by('brand', 'name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        products = context['products']
        all_variants = ProductVariant.objects.filter(product__in=products).select_related('product').order_by('product__brand', 'product__name', 'sku')
        context['variants'] = all_variants
        return context


class BarcodeGeneratePDFView(CashierRequiredMixin, View):
    """Generate PDF with barcodes for selected variants"""
    def post(self, request):
        import io
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.pdfgen import canvas
        from reportlab.lib.utils import ImageReader
        from django.http import HttpResponse

        try:
            import barcode
            from barcode.writer import ImageWriter
        except ImportError:
            messages.error(request, 'Barcode library not installed. Please install python-barcode.')
            return redirect('barcode_generator')

        variant_ids = request.POST.getlist('variants')
        if not variant_ids:
            messages.error(request, 'Please select at least one item to generate barcodes.')
            return redirect('barcode_generator')

        barcode_mode = request.POST.get('barcode_mode', 'single')
        
        variants = ProductVariant.objects.filter(id__in=variant_ids).select_related('product')
        
        barcodes_to_generate = []
        for variant in variants:
            if barcode_mode == 'quantity':
                quantity = variant.initial_quantity
                for i in range(max(1, quantity)):                                  
                    barcodes_to_generate.append(variant)
            else:
                barcodes_to_generate.append(variant)
        
        buffer = io.BytesIO()
        p = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4
        
        barcode_width = 60 * mm
        barcode_height = 25 * mm
        label_height = 40 * mm                               
        margin_x = 15 * mm
        margin_y = 15 * mm
        spacing_x = 5 * mm
        spacing_y = 8 * mm
        
        cols = 3
        col = 0
        row = 0
        max_rows = int((height - 2 * margin_y) / (label_height + spacing_y))
        
        def draw_barcode(p, variant, x, y):
            """Helper function to draw a single barcode"""
            try:
                CODE128 = barcode.get_barcode_class('code128')
                code = CODE128(variant.sku, writer=ImageWriter())
                barcode_buffer = io.BytesIO()
                code.write(barcode_buffer, options={
                    'module_width': 0.3,
                    'module_height': 8,
                    'font_size': 0,
                    'text_distance': 1,
                    'quiet_zone': 2,
                })
                barcode_buffer.seek(0)
                barcode_img = ImageReader(barcode_buffer)
                
                product_name = f"{variant.product.brand} {variant.product.name}"
                if len(product_name) > 28:
                    product_name = product_name[:25] + "..."
                
                p.setFont("Helvetica-Bold", 8)
                p.drawString(x, y + label_height - 8, product_name)
                
                variant_info = f"{variant.color or ''} / {variant.size or ''}"
                p.setFont("Helvetica", 7)
                p.drawString(x, y + label_height - 18, variant_info.strip(' /'))
                
                p.drawImage(barcode_img, x, y + 8, width=barcode_width, height=barcode_height - 5, preserveAspectRatio=True)
                
                p.setFont("Helvetica", 8)
                p.drawCentredString(x + barcode_width / 2, y, variant.sku)
                
            except Exception as e:
                p.setFont("Helvetica-Bold", 9)
                product_display = f"{variant.product.brand} {variant.product.name}"[:30]
                p.drawString(x, y + 20, product_display)
                p.setFont("Helvetica", 8)
                p.drawString(x, y + 8, f"SKU: {variant.sku}")
        
        for variant in barcodes_to_generate:
            x = margin_x + col * (barcode_width + spacing_x)
            y = height - margin_y - (row + 1) * (label_height + spacing_y)
            
            draw_barcode(p, variant, x, y)
            
            col += 1
            if col >= cols:
                col = 0
                row += 1
                if row >= max_rows:
                    p.showPage()
                    row = 0
        
        p.save()
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="barcodes.pdf"'
        return response


class ProductDetailView(LoginRequiredMixin, DetailView):
    model = Product
    template_name = 'core/product_detail.html'
    context_object_name = 'product'

class QuickProductUpdateView(CashierRequiredMixin, UpdateView):
    model = Product
    form_class = ProductForm
    template_name = 'core/product_form.html'
    success_url = reverse_lazy('product_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['editing'] = True
        context['brand_suggestions'] = BRAND_SUGGESTIONS
        return context

class LuxuryProductUpdateView(OwnerRequiredMixin, View):
    template_name = 'core/luxury_product_form.html'
    success_url = reverse_lazy('product_list')

    def _normalize_image(self, image_field):
        if not image_field:
            return
        path = image_field.path
        try:
            img = Image.open(path).convert('RGB')
        except Exception:
            return
        max_side = 1400
        img.thumbnail((max_side, max_side), Image.LANCZOS)
        canvas_size = max(img.size)
        canvas = Image.new('RGB', (canvas_size, canvas_size), (255, 255, 255))
        offset = ((canvas_size - img.size[0]) // 2, (canvas_size - img.size[1]) // 2)
        canvas.paste(img, offset)
        canvas.save(path, format='JPEG', quality=90)

    def get(self, request, pk):
        product = get_object_or_404(Product, pk=pk)
        product_form = LuxuryProductForm(instance=product, user=request.user)
        variant_formset = ProductVariantFormSet(
            instance=product,
            prefix='variants',
            form_kwargs={'user': request.user}
        )
        context = {
            'form': product_form,
            'variant_formset': variant_formset,
            'user': request.user,
            'brand_suggestions': BRAND_SUGGESTIONS,
            'material_choices': [choice[0] for choice in MATERIAL_CHOICES],
            'editing': True,
            'product': product,
        }
        return render(request, self.template_name, context)

    def post(self, request, pk):
        product = get_object_or_404(Product, pk=pk)
        product_form = LuxuryProductForm(request.POST, request.FILES, instance=product, user=request.user)
        variant_formset = ProductVariantFormSet(
            request.POST,
            instance=product,
            prefix='variants',
            form_kwargs={'user': request.user}
        )
        additional_images = request.FILES.getlist('additional_images')
        
        if product_form.is_valid() and variant_formset.is_valid():
            with transaction.atomic():
                product = product_form.save()
                self._normalize_image(product.image)
                
                variants = variant_formset.save(commit=False)
                for variant in variants:
                    variant.product = product
                    if variant.cost_price is None:
                        variant.cost_price = Decimal('0.00')
                    if not variant.currency:
                        variant.currency = ProductVariant.CURRENCY_CHOICES[0][0]
                    if not variant.sku:
                        color_part = (variant.color or 'CLR')[:3].upper().replace(' ', '')
                        size_part = (variant.size or 'SZ')[:3].upper().replace(' ', '')
                        base_sku = product.base_sku
                        candidate = f"{base_sku}-{color_part}-{size_part}"
                        counter = 1
                        while ProductVariant.objects.filter(sku=candidate).exclude(id=variant.id).exists():
                            candidate = f"{base_sku}-{color_part}-{size_part}-{counter}"
                            counter += 1
                        variant.sku = candidate
                    variant.save()
                    
                
                for vform in variant_formset.forms:
                    if vform.cleaned_data.get('DELETE'):
                        continue
                    if vform.instance.pk:               
                        code = vform.cleaned_data.get('barcode')
                        if code:
                            Barcode.objects.update_or_create(
                                variant=vform.instance,
                                defaults={
                                    'code': code,
                                    'barcode_type': vform.cleaned_data.get('barcode_type') or Barcode.Type.INTERNAL
                                }
                            )

                for obj in variant_formset.deleted_objects:
                    obj.delete()

                start_order = product.additional_images.count()
                for idx, image in enumerate(additional_images):
                    ProductImage.objects.create(
                        product=product,
                        image=image,
                        order=start_order + idx,
                        is_primary=False
                    )
                
                messages.success(request, "Luxury product updated successfully.")
                return redirect(self.success_url)
        
        context = {
            'form': product_form,
            'variant_formset': variant_formset,
            'user': request.user,
            'brand_suggestions': BRAND_SUGGESTIONS,
            'material_choices': [choice[0] for choice in MATERIAL_CHOICES],
            'editing': True,
            'product': product,
        }
        return render(request, self.template_name, context)

class ProductUpdateView(CashierRequiredMixin, View):
    def dispatch(self, request, *args, **kwargs):
        pk = kwargs.get('pk')
        product = get_object_or_404(Product, pk=pk)
        if product.is_luxury_product:
            return LuxuryProductUpdateView.as_view()(request, *args, **kwargs)
        else:
            return QuickProductUpdateView.as_view()(request, *args, **kwargs)

class ProductDeleteView(OwnerRequiredMixin, View):
    template_name = 'core/product_confirm_delete.html'
    success_url = reverse_lazy('product_list')

    def get(self, request, pk):
        product = get_object_or_404(Product, pk=pk)
        return render(request, self.template_name, {'product': product})

    def post(self, request, pk):
        product = get_object_or_404(Product, pk=pk)
        try:
            product.delete()
            messages.success(request, "Product deleted.")
        except ProtectedError:
            messages.error(request, "Cannot delete this product because it has been sold or referenced in other records. Consider archiving it instead.")
        return redirect(self.success_url)

class ProductCreateView(CashierRequiredMixin, View):
    """Quick add product flow focused on required fields and fast variant capture"""
    template_name = 'core/product_form.html'
    success_url = reverse_lazy('product_list')

    def _normalize_image(self, image_field):
        try:
            if not image_field:
                return
            path = image_field.path
            img = Image.open(path)
            img_format = img.format or 'JPEG'
            img = img.convert('RGB')
            max_side = 1400
            img.thumbnail((max_side, max_side), Image.LANCZOS)
            canvas_size = max(img.size)
            canvas = Image.new('RGB', (canvas_size, canvas_size), (255, 255, 255))
            offset = ((canvas_size - img.size[0]) // 2, (canvas_size - img.size[1]) // 2)
            canvas.paste(img, offset)
            canvas.save(path, format=img_format, quality=90)
        except Exception:
            return

    def get(self, request):
        form = ProductForm()
        return render(
            request,
            self.template_name,
            {
                'form': form,
                'brand_suggestions': BRAND_SUGGESTIONS,
            }
        )

    def post(self, request):
        form = ProductForm(request.POST, request.FILES)

        variant_rows = []
        try:
            total = int(request.POST.get('quick_variants-TOTAL_FORMS', 0))
        except (TypeError, ValueError):
            total = 0
        invalid_variant_data = False
        for i in range(total):
            color = request.POST.get(f'quick_variants-{i}-color', '').strip()
            size = request.POST.get(f'quick_variants-{i}-size', '').strip()
            qty = request.POST.get(f'quick_variants-{i}-quantity')
            cost = request.POST.get(f'quick_variants-{i}-cost_price')
            retail = request.POST.get(f'quick_variants-{i}-retail_price')
            if not qty and not retail and not color and not size and not cost:
                continue
            try:
                qty_val = int(qty)
            except (TypeError, ValueError):
                qty_val = None
            try:
                cost_val = Decimal(cost) if cost else None
            except (TypeError, ValueError, InvalidOperation):
                cost_val = None
            try:
                retail_val = Decimal(retail)
            except (TypeError, ValueError, InvalidOperation):
                retail_val = None

            variant_rows.append({
                'color': color or 'Standard',
                'size': size or 'One Size',
                'quantity': qty_val,
                'cost_price': cost_val,
                'retail_price': retail_val,
            })
            if qty_val is None or qty_val < 1 or cost_val is None or cost_val < 0 or retail_val is None or retail_val < 0:
                invalid_variant_data = True

        seen_variants = set()
        for row in variant_rows:
            key = (row['color'].lower(), row['size'].lower())
            if key in seen_variants:
                invalid_variant_data = True
                messages.error(request, f"Duplicate variant found: {row['color']} / {row['size']}. Please combine quantities.")
            seen_variants.add(key)

        if not variant_rows:
            messages.error(request, "Add at least one color/size row with quantity.")
        if invalid_variant_data and not messages.get_messages(request):
            messages.error(request, "Enter valid quantities, cost prices, and retail prices for each row.")

        if form.is_valid() and variant_rows and not invalid_variant_data:
            with transaction.atomic():
                product = form.save()
                self._normalize_image(product.image)
                for idx, row in enumerate(variant_rows):
                    color_part = (row['color'] or 'COL')[:3].upper().replace(' ', '')
                    size_part = (row['size'] or 'SZ')[:3].upper().replace(' ', '')
                    base_sku_val = f"{product.base_sku}-{color_part}-{size_part}"
                    
                    sku_value = base_sku_val
                    counter = 1
                    while ProductVariant.objects.filter(sku=sku_value).exists():
                        sku_value = f"{base_sku_val}-{counter}"
                        counter += 1

                    ProductVariant.objects.create(
                        product=product,
                        sku=sku_value,
                        color=row['color'],
                        size=row['size'],
                        hardware_type='',
                        variant_material='',
                        cost_price=row['cost_price'] or Decimal('0.00'),
                        retail_price=row['retail_price'] or Decimal('0.00'),
                        currency=ProductVariant.CURRENCY_CHOICES[0][0],
                        serial_number=None,
                        authentication_code='',
                        edition_number='',
                        initial_quantity=row['quantity'] or 0,
                        minimum_stock_level=1,
                    )
                messages.success(request, "Product added quickly with starter variants.")
                return redirect(self.success_url)

        return render(
            request,
            self.template_name,
            {
                'form': form,
                'quick_variant_rows': variant_rows,
                'brand_suggestions': BRAND_SUGGESTIONS,
            }
        )

class LuxuryProductCreateView(CashierRequiredMixin, CreateView):
    """Comprehensive luxury product registration view with variant + media support"""
    template_name = 'core/luxury_product_form.html'
    success_url = reverse_lazy('product_list')

    def _normalize_image(self, image_field):
        if not image_field:
            return
        path = image_field.path
        try:
            img = Image.open(path).convert('RGB')
        except Exception:
            return
        max_side = 1400
        img.thumbnail((max_side, max_side), Image.LANCZOS)
        canvas_size = max(img.size)
        canvas = Image.new('RGB', (canvas_size, canvas_size), (255, 255, 255))
        offset = ((canvas_size - img.size[0]) // 2, (canvas_size - img.size[1]) // 2)
        canvas.paste(img, offset)
        canvas.save(path, format='JPEG', quality=90)

    def get(self, request):
        product_form = LuxuryProductForm(user=request.user)
        variant_formset = ProductVariantFormSet(
            prefix='variants',
            instance=Product(),
            form_kwargs={'user': request.user}
        )
        context = {
            'form': product_form,
            'variant_formset': variant_formset,
            'user': request.user,
            'brand_suggestions': BRAND_SUGGESTIONS,
            'material_choices': [choice[0] for choice in MATERIAL_CHOICES],
            'color_suggestions': COLOR_SUGGESTIONS,
        }
        return render(request, self.template_name, context)

    def post(self, request):
        product_form = LuxuryProductForm(request.POST, request.FILES, user=request.user)
        variant_formset = ProductVariantFormSet(
            request.POST,
            prefix='variants',
            instance=Product(),
            form_kwargs={'user': request.user}
        )
        additional_images = request.FILES.getlist('additional_images')
        
        if product_form.is_valid() and variant_formset.is_valid():
            with transaction.atomic():
                product = product_form.save(commit=False)
                product.is_luxury_product = True
                product.save()
                product_form.save_m2m()
                self._normalize_image(product.image)
                
                saved_variants = []
                for vform in variant_formset.forms:
                    if vform.cleaned_data.get('DELETE'):
                        continue
                    if not vform.has_changed() and not vform.instance.pk:
                        continue
                    variant = vform.save(commit=False)
                    variant.product = product
                    if variant.cost_price is None:
                        variant.cost_price = Decimal('0.00')
                    if not variant.currency:
                        variant.currency = ProductVariant.CURRENCY_CHOICES[0][0]
                    if not variant.sku:
                        color_part = (variant.color or 'CLR')[:3].upper().replace(' ', '')
                        size_part = (variant.size or 'SZ')[:3].upper().replace(' ', '')
                        base_sku = product.base_sku
                        candidate = f"{base_sku}-{color_part}-{size_part}"
                        counter = 1
                        while ProductVariant.objects.filter(sku=candidate).exists():
                            candidate = f"{base_sku}-{color_part}-{size_part}-{counter}"
                            counter += 1
                        variant.sku = candidate
                    variant.save()
                    saved_variants.append(variant)
                    
                    code = vform.cleaned_data.get('barcode')
                    if code:
                        Barcode.objects.update_or_create(
                            variant=variant,
                            defaults={
                                'code': code,
                                'barcode_type': vform.cleaned_data.get('barcode_type') or Barcode.Type.INTERNAL
                            }
                        )
                
                start_order = 0
                for idx, image in enumerate(additional_images):
                    ProductImage.objects.create(
                        product=product,
                        image=image,
                        order=start_order + idx,
                        is_primary=False
                    )
                
                messages.success(request, "Luxury product registered successfully.")
                return redirect(self.success_url)
        
        context = {
            'form': product_form,
            'variant_formset': variant_formset,
            'user': request.user,
            'brand_suggestions': BRAND_SUGGESTIONS,
            'material_choices': [choice[0] for choice in MATERIAL_CHOICES],
        }
        return render(request, self.template_name, context)

class ClientListView(SalesAssociateRequiredMixin, ListView):
    model = Client
    template_name = 'core/client_list.html'
    context_object_name = 'clients'
    paginate_by = 20

    def get_queryset(self):
        queryset = super().get_queryset().order_by('first_name', 'last_name')
        
        q = self.request.GET.get('q', '').strip()
        if q:
            queryset = queryset.filter(
                Q(first_name__icontains=q) |
                Q(last_name__icontains=q) |
                Q(phone__icontains=q) |
                Q(phone__endswith=q) |                 
                Q(email__icontains=q)
            ).distinct()
        
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('q', '')
        return context

class ClientCreateView(CashierRequiredMixin, CreateView):
    model = Client
    form_class = ClientForm
    template_name = 'core/client_form.html'
    success_url = reverse_lazy('client_list')

class ClientUpdateView(CashierRequiredMixin, UpdateView):
    model = Client
    form_class = ClientForm
    template_name = 'core/client_form.html'
    success_url = reverse_lazy('client_list')

class ClientDetailView(SalesAssociateRequiredMixin, DetailView):
    model = Client
    template_name = 'core/client_detail.html'
    context_object_name = 'client'

class CashierOnlyMixin(LoginRequiredMixin):
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated or not request.user.is_sales_associate():
            return self.handle_no_permission()
        return super().dispatch(request, *args, **kwargs)

class POSView(CashierRequiredMixin, TemplateView):
    template_name = 'core/pos.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        clients = Client.objects.all().order_by('first_name')
        variants = ProductVariant.objects.select_related('product').all()
        variant_data = [
            {
                'id': v.id,
                'sku': v.sku,
                'name': f"{v.product.brand} {v.product.name}",
                'color': v.color,
                'size': v.size,
                'price': float(v.retail_price),
                'stock': v.initial_quantity,                         
                'image': v.product.image.url if v.product.image else '',
            }
            for v in variants
        ]
        context.update({
            'clients': clients,
            'variant_data': variant_data,
        })
        return context


@method_decorator(csrf_exempt, name='dispatch')
class POSPreviewDiscountView(CashierOnlyMixin, View):
    """Preview discount calculation without creating an order"""
    def post(self, request):
        try:
            payload = json.loads(request.body.decode('utf-8'))
        except Exception:
            return JsonResponse({'error': 'Invalid payload'}, status=400)
        
        items = payload.get('items') or []
        client_id = payload.get('client_id')
        promo_code = payload.get('promo_code', '').strip().upper() if payload.get('promo_code') else None
        
        if not items:
            return JsonResponse({'discount': 0, 'description': ''})
        
        client_obj = None
        if client_id:
            try:
                client_obj = Client.objects.get(id=client_id)
            except Client.DoesNotExist:
                pass
        
        cart_items = []
        subtotal = Decimal('0.00')
        for item in items:
            sku = item.get('sku')
            qty = int(item.get('qty') or 0)
            price = Decimal(str(item.get('price') or '0'))
            if not sku or qty <= 0:
                continue
            try:
                variant = ProductVariant.objects.select_related('product').get(sku=sku)
                cart_items.append({
                    'sku': sku,
                    'qty': qty,
                    'price': float(price),
                    'total': float(price * qty),
                    'product': variant.product,
                })
                subtotal += price * qty
            except ProductVariant.DoesNotExist:
                continue
        
        if not cart_items:
            return JsonResponse({'discount': 0, 'description': ''})
        
        best_discount = Decimal('0.00')
        best_desc = ''
        
        if promo_code:
            try:
                code_promo = Promotion.objects.get(
                    code__iexact=promo_code,
                    is_active=True,
                    start_date__lte=timezone.now(),
                    end_date__gte=timezone.now()
                )
                if code_promo.max_uses == 0 or code_promo.used_count < code_promo.max_uses:
                    promo_amount, promo_desc = code_promo.calculate_discount(cart_items, client_obj)
                    if promo_amount > 0:
                        best_discount = promo_amount
                        best_desc = f"{code_promo.name}: {promo_desc}"
            except Promotion.DoesNotExist:
                return JsonResponse({
                    'discount': 0, 
                    'description': '',
                    'error': 'Invalid promo code'
                })
        
        if best_discount == 0:
            active_promos = Promotion.objects.filter(
                is_active=True,
                start_date__lte=timezone.now(),
                end_date__gte=timezone.now(),
            ).filter(Q(code__isnull=True) | Q(code=''))
            
            for promo in active_promos:
                if promo.max_uses > 0 and promo.used_count >= promo.max_uses:
                    continue
                promo_amount, promo_desc = promo.calculate_discount(cart_items, client_obj)
                if promo_amount > best_discount:
                    best_discount = promo_amount
                    best_desc = f"{promo.name}: {promo_desc}"
        
        return JsonResponse({
            'discount': float(best_discount),
            'description': best_desc,
            'subtotal': float(subtotal),
            'total': float(subtotal - best_discount)
        })

@method_decorator(csrf_exempt, name='dispatch')
class POSCheckoutView(CashierOnlyMixin, View):
    """Create an order from POS items and return order details for receipt"""
    def _generate_code(self):
        while True:
            code = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))
            if not Order.objects.filter(order_code=code).exists():
                return code

    def post(self, request):
        import json
        try:
            payload = json.loads(request.body.decode('utf-8'))
        except Exception:
            return HttpResponseBadRequest("Invalid payload")
        
        items = payload.get('items') or []
        client_name = payload.get('client') or ''
        client_id = payload.get('client_id')
        promo_code = payload.get('promo_code', '').strip().upper() if payload.get('promo_code') else None
        client_obj = None
        if client_id:
            try:
                client_obj = Client.objects.get(id=client_id)
                client_name = f"{client_obj.first_name} {client_obj.last_name}".strip() or client_obj.phone or client_name
            except Client.DoesNotExist:
                client_obj = None
        discount = Decimal(payload.get('discount') or 0)
        if not items:
            return HttpResponseBadRequest("No items provided")
        
        for item in items:
            sku = item.get('sku')
            qty = int(item.get('qty') or 0)
            if not sku or qty <= 0:
                continue
            try:
                variant = ProductVariant.objects.get(sku=sku)
                if variant.initial_quantity < qty:
                    product_name = f"{variant.product.brand} {variant.product.name}"
                    variant_desc = f"{variant.color}/{variant.size}" if variant.color or variant.size else "variant"
                    return JsonResponse({
                        'error': f'Insufficient stock for {product_name} ({variant_desc}). Only {variant.initial_quantity} available, but {qty} requested.'
                    }, status=400)
            except ProductVariant.DoesNotExist:
                return JsonResponse({'error': f'Product with SKU {sku} not found'}, status=400)
        
        location = Location.objects.first()
        if not location:
            location = Location.objects.create(name='Main Warehouse', code='MAIN', address='N/A')
        
        with transaction.atomic():
            order = Order.objects.create(
                order_code=self._generate_code(),
                client=client_obj,
                location=location,
                created_by=request.user,
                status=Order.Status.COMPLETED,
                total_amount=Decimal('0.00'),
                total_discount=discount,
            )
            total = Decimal('0.00')
            created_items = []
            for item in items:
                sku = item.get('sku')
                qty = int(item.get('qty') or 0)
                price = Decimal(str(item.get('price') or '0'))
                if not sku or qty <= 0:
                    continue
                try:
                    variant = ProductVariant.objects.get(sku=sku)
                except ProductVariant.DoesNotExist:
                    continue
                line_total = price * qty
                total += line_total
                oi = OrderItem.objects.create(
                    order=order,
                    variant=variant,
                    quantity=qty,
                    unit_price=price,
                    line_discount=Decimal('0.00'),
                )
                variant.initial_quantity = max(0, variant.initial_quantity - qty)
                variant.save(update_fields=['initial_quantity'])
                created_items.append({
                    'name': f"{variant.product.brand} {variant.product.name}",
                    'sku': variant.sku,
                    'color': variant.color,
                    'size': variant.size,
                    'qty': qty,
                    'price': float(price),
                    'total': float(line_total),
                    'product': variant.product,                             
                })
            
            promo_discount = Decimal('0.00')
            promo_description = ''
            
            best_discount = Decimal('0.00')
            best_promo = None
            best_desc = ''
            
            if promo_code:
                try:
                    code_promo = Promotion.objects.get(
                        code__iexact=promo_code,
                        is_active=True,
                        start_date__lte=timezone.now(),
                        end_date__gte=timezone.now()
                    )
                    if code_promo.max_uses == 0 or code_promo.used_count < code_promo.max_uses:
                        promo_amount, promo_desc = code_promo.calculate_discount(created_items, client_obj)
                        if promo_amount > 0:
                            best_discount = promo_amount
                            best_promo = code_promo
                            best_desc = promo_desc
                except Promotion.DoesNotExist:
                    pass                                                
            
            if not best_promo:
                active_promos = Promotion.objects.filter(
                    is_active=True,
                    start_date__lte=timezone.now(),
                    end_date__gte=timezone.now(),
                    code__isnull=True                                        
                ) | Promotion.objects.filter(
                    is_active=True,
                    start_date__lte=timezone.now(),
                    end_date__gte=timezone.now(),
                    code=''
                )
                
                for promo in active_promos:
                    if promo.max_uses > 0 and promo.used_count >= promo.max_uses:
                        continue
                    promo_amount, promo_desc = promo.calculate_discount(created_items, client_obj)
                    if promo_amount > best_discount:
                        best_discount = promo_amount
                        best_promo = promo
                        best_desc = promo_desc
            
            if best_promo and best_discount > 0:
                promo_discount = best_discount
                promo_description = f"{best_promo.name}: {best_desc}"
                best_promo.used_count += 1
                best_promo.save(update_fields=['used_count'])
                PromotionUsage.objects.create(
                    promotion=best_promo,
                    client=client_obj,
                    order=order,
                    discount_amount=promo_discount
                )
            
            total_discount = discount + promo_discount
            order.total_amount = total - total_discount
            order.total_discount = total_discount
            order.save(update_fields=['total_amount', 'total_discount'])
        
        response_items = [
            {k: v for k, v in item.items() if k != 'product'}
            for item in created_items
        ]
        
        return JsonResponse({
            'order_id': order.id,
            'order_code': order.order_code,
            'client': client_name or 'Walk-in',
            'cashier': request.user.username,
            'total': float(order.total_amount),
            'discount': float(order.total_discount),
            'subtotal': float(total),
            'items': response_items,
            'created_at': order.created_at.isoformat(),
        })

def calculate_order_return_status(order):
    """
    Calculate if SALE/EXCHANGE order is COMPLETED, PARTIALLY_RETURNED, or FULLY_RETURNED
    based on item-level qty_returned tracking
    """
    if order.type not in [Order.Type.SALE, Order.Type.EXCHANGE]:
        return order.status
    
    items = order.items.all()
    if not items:
        return Order.Status.COMPLETED
    
    all_fully_returned = all(item.is_fully_returned for item in items)
    any_partially_returned = any(item.qty_returned > 0 for item in items)
    
    if all_fully_returned:
        return Order.Status.FULLY_RETURNED
    elif any_partially_returned:
        return Order.Status.PARTIALLY_RETURNED
    else:
        return Order.Status.COMPLETED

@method_decorator(csrf_exempt, name='dispatch')
class POSReturnLookupView(CashierOnlyMixin, View):
    def post(self, request):
        import json
        try:
            payload = json.loads(request.body.decode('utf-8'))
            code = payload.get('order_code', '').strip().upper().replace('#', '')
        except Exception:
            return JsonResponse({'error': 'Invalid payload'}, status=400)
        if not code:
            return JsonResponse({'error': 'Order code required'}, status=400)
        try:
            order = Order.objects.get(order_code=code)
        except Order.DoesNotExist:
            if code.isdigit():
                try:
                    order = Order.objects.get(id=int(code))
                except Order.DoesNotExist:
                    return JsonResponse({'error': f'Order not found with code "{code}"'}, status=404)
            else:
                return JsonResponse({'error': f'Order not found with code "{code}"'}, status=404)
        
        if order.type == Order.Type.RETURN:
            return JsonResponse({
                'allow_return': False,
                'reason': 'This is a return/refund order and cannot be returned again.'
            })
        
        if order.status == Order.Status.FULLY_RETURNED:
            return JsonResponse({
                'allow_return': False,
                'reason': 'This order has already been fully returned. No items are available for return or exchange.'
            })

        return_window_days = 10
        days_since_purchase = (timezone.now() - order.created_at).days
        within_return_window = days_since_purchase <= return_window_days

        items = []
        has_returnable_items = False
        for oi in order.items.select_related('variant__product').all():
            qty_remaining = oi.qty_remaining
            
            if qty_remaining == 0:
                status = 'already_returned'
            elif not within_return_window:
                status = 'not_eligible'
            elif qty_remaining > 0:
                status = 'returnable'
                has_returnable_items = True
            else:
                status = 'already_returned'
            
            items.append({
                'id': oi.id,
                'sku': oi.variant.sku,
                'name': f"{oi.variant.product.brand} {oi.variant.product.name}",
                'color': oi.variant.color,
                'size': oi.variant.size,
                'qty': qty_remaining,                       
                'original_qty': oi.quantity,
                'returned_qty': oi.qty_returned,
                'price': float(oi.unit_price),
                'status': status,
            })
        
        if not has_returnable_items:
            if any(item['status'] == 'already_returned' for item in items):
                return JsonResponse({
                    'allow_return': False,
                    'reason': 'This order has already been fully returned. No items are available for return or exchange.'
                })
            return JsonResponse({
                'allow_return': False,
                'reason': 'No items are eligible for return. Return window has expired.'
            })

        return JsonResponse({
            'order_id': order.id,
            'order_code': order.order_code,
            'client': str(order.client) if order.client else 'Walk-in',
            'items': items,
            'created_at': order.created_at.isoformat(),
            'allow_return': True,
            'days_since_purchase': days_since_purchase,
            'return_window_days': return_window_days,
        })

@method_decorator(csrf_exempt, name='dispatch')
class POSReturnCheckoutView(CashierRequiredMixin, View):
    def _gen_code(self):
        while True:
            code = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))
            if not Order.objects.filter(order_code=code).exists():
                return code

    def post(self, request):
        import json
        try:
            payload = json.loads(request.body.decode('utf-8'))
        except Exception as e:
            return JsonResponse({'error': f'Invalid payload: {str(e)}'}, status=400)

        order_code = payload.get('order_code')
        reason = payload.get('reason', 'OTHER')
        action = payload.get('action', 'REFUND')
        return_items = payload.get('return_items') or []
        replacement_items = payload.get('replacement_items') or []

        if not order_code:
            return JsonResponse({'error': 'Order code is required'}, status=400)

        try:
            original_order = Order.objects.get(order_code=order_code)
        except Order.DoesNotExist:
            if order_code.isdigit():
                try:
                    original_order = Order.objects.get(id=int(order_code))
                except Order.DoesNotExist:
                    return JsonResponse({'error': f'Order not found: {order_code}'}, status=404)
            else:
                return JsonResponse({'error': f'Order not found: {order_code}'}, status=404)

        location = original_order.location
        client_obj = original_order.client

        try:
            with transaction.atomic():
                return_obj = Return.objects.create(
                    original_order=original_order,
                    created_by=request.user,
                    reason=reason,
                    action=action
                )

                refund_amount = Decimal('0.00')
                
                returned_map = {}
                for ret in original_order.returns.all():
                    for ri in ret.items.all():
                        returned_map[ri.order_item_id] = returned_map.get(ri.order_item_id, 0) + ri.quantity

                for item in replacement_items:
                    sku = item.get('sku')
                    qty = int(item.get('qty') or 0)
                    if not sku or qty <= 0:
                        continue
                    try:
                        variant = ProductVariant.objects.get(sku=sku)
                        if variant.initial_quantity < qty:
                            product_name = f"{variant.product.brand} {variant.product.name}"
                            variant_desc = f"{variant.color}/{variant.size}" if variant.color or variant.size else "variant"
                            return JsonResponse({
                                'error': f'Insufficient stock for exchange item: {product_name} ({variant_desc}). Only {variant.initial_quantity} available, but {qty} requested.'
                            }, status=400)
                    except ProductVariant.DoesNotExist:
                        return JsonResponse({'error': f'Product with SKU {sku} not found'}, status=400)

                for item in return_items:
                    order_item_id = item.get('id')
                    qty = int(item.get('qty') or 0)
                    if qty <= 0:
                        continue
                    
                    try:
                        order_item = OrderItem.objects.get(id=order_item_id, order=original_order)
                    except OrderItem.DoesNotExist:
                        continue
                    
                    if qty > order_item.qty_remaining:
                        raise Exception(f"Cannot return {qty} of {order_item.variant}. Only {order_item.qty_remaining} remaining (already returned {order_item.qty_returned}).")

                    ReturnItem.objects.create(
                        return_process=return_obj,
                        order_item=order_item,
                        quantity=qty,
                        reason=reason
                    )

                    order_item.qty_returned = F('qty_returned') + qty
                    order_item.save(update_fields=['qty_returned'])
                    order_item.refresh_from_db()                     

                    variant = order_item.variant
                    variant.initial_quantity = variant.initial_quantity + qty
                    variant.save(update_fields=['initial_quantity'])

                    refund_amount += order_item.unit_price * qty

                replacement_total = Decimal('0.00')
                replacement_lines = []
                
                for item in replacement_items:
                    sku = item.get('sku')
                    qty = int(item.get('qty') or 0)
                    price = Decimal(str(item.get('price') or '0'))
                    if not sku or qty <= 0:
                        continue
                    try:
                        variant = ProductVariant.objects.get(sku=sku)
                    except ProductVariant.DoesNotExist:
                        continue
                    
                    replacement_total += price * qty
                    variant.initial_quantity = max(0, variant.initial_quantity - qty)
                    variant.save(update_fields=['initial_quantity'])
                    replacement_lines.append({
                        'variant': variant,
                        'qty': qty,
                        'price': price,
                    })

                
                replacement_order = None
                if replacement_lines:
                    replacement_order = Order.objects.create(
                        type=Order.Type.EXCHANGE,                         
                        parent_order=original_order,                    
                        order_code=self._gen_code(),
                        client=client_obj,
                        location=location,
                        created_by=request.user,
                        status=Order.Status.COMPLETED,
                        total_amount=replacement_total,
                        total_discount=Decimal('0.00'),
                    )
                    
                    for line in replacement_lines:
                        OrderItem.objects.create(
                            order=replacement_order,
                            variant=line['variant'],
                            quantity=line['qty'],
                            unit_price=line['price'],
                            line_discount=Decimal('0.00'),
                        )
                    
                    return_obj.replacement_order = replacement_order
                
                return_obj.save()
                
                original_order.status = calculate_order_return_status(original_order)
                original_order.save(update_fields=['status'])

                returned_items_list = []
                for item in return_items:
                    try:
                        oi = OrderItem.objects.get(id=item.get('id'))
                        returned_items_list.append({
                            'name': f"{oi.variant.product.brand} {oi.variant.product.name}",
                            'sku': oi.variant.sku,
                            'color': oi.variant.color or '',
                            'size': oi.variant.size or '',
                            'qty': int(item.get('qty')),
                            'price': float(oi.unit_price),
                            'total': float(oi.unit_price * int(item.get('qty')))
                        })
                    except: 
                        pass

                replacement_items_list = []
                for line in replacement_lines:
                    v = line['variant']
                    replacement_items_list.append({
                        'name': f"{v.product.brand} {v.product.name}",
                        'sku': v.sku,
                        'color': v.color or '',
                        'size': v.size or '',
                        'qty': line['qty'],
                        'price': float(line['price']),
                        'total': float(line['price'] * line['qty'])
                    })

                net = float(replacement_total) - float(refund_amount)

                return JsonResponse({
                    'success': True,
                    'return_id': return_obj.id,
                    'original_order_code': original_order.order_code,
                    'original_order_status': original_order.status,
                    'refund_amount': float(refund_amount),
                    'exchange_order_code': replacement_order.order_code if replacement_order else None,
                    'exchange_total': float(replacement_total) if replacement_order else 0,
                    'net_due': net,
                    'action': action,
                    'reason': reason,
                    'returned_items': returned_items_list,
                    'replacement_items': replacement_items_list,
                    'created_at': timezone.now().isoformat(),
                    'cashier': request.user.username,
                    'client': str(client_obj) if client_obj else 'Walk-in',
                })
        except Exception as e:
            return JsonResponse({'error': f'Error processing return: {str(e)}'}, status=400)

class POSReturnView(CashierRequiredMixin, TemplateView):
    template_name = 'core/pos_return.html'

class ReportView(OwnerRequiredMixin, TemplateView):
    template_name = 'core/reports.html'
    
    def get_context_data(self, **kwargs):
        from datetime import timedelta
        from django.db.models.functions import TruncDate
        
        context = super().get_context_data(**kwargs)
        
        days = int(self.request.GET.get('days', 30))
        end_date = timezone.now()
        start_date = end_date - timedelta(days=days)
        
        context['selected_days'] = days
        
        completed_orders = Order.objects.filter(
            status=Order.Status.COMPLETED,
            created_at__range=(start_date, end_date)
        )
        
        refunded_orders = Order.objects.filter(
            created_at__range=(start_date, end_date)
        ).filter(Q(status=Order.Status.REFUNDED) | Q(type=Order.Type.RETURN))
        
        total_revenue = completed_orders.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')

        total_refunds = Decimal('0')
        returns_qs = Return.objects.filter(created_at__range=(start_date, end_date)).prefetch_related('items__order_item')
        for ret in returns_qs:
            for item in ret.items.all():
                if item.order_item and item.order_item.unit_price:
                    total_refunds += item.order_item.unit_price * item.quantity

        net_revenue = total_revenue - total_refunds
        
        sales_cost = Decimal('0')
        for order in completed_orders:
            for item in order.items.all():
                if item.variant and item.variant.cost_price:
                    sales_cost += item.variant.cost_price * item.quantity
        
        return_cost = Decimal('0')
        for ret in returns_qs:
            for item in ret.items.all():
                if item.order_item and item.order_item.variant and item.order_item.variant.cost_price:
                    return_cost += item.order_item.variant.cost_price * item.quantity

        total_cost = sales_cost - return_cost
        
        gross_profit = total_revenue - sales_cost
        gross_margin = (gross_profit / total_revenue * 100) if total_revenue else Decimal('0')

        total_profit = net_revenue - total_cost
        margin_base = total_revenue if total_revenue > 0 else net_revenue
        profit_margin = (total_profit / margin_base * 100) if margin_base and margin_base != 0 else Decimal('0')
        
        total_orders = completed_orders.count()
        total_items = OrderItem.objects.filter(order__in=completed_orders).aggregate(total=Sum('quantity'))['total'] or 0
        avg_order_value = net_revenue / total_orders if total_orders > 0 else Decimal('0')
        
        daily_sales = completed_orders.annotate(
            date=TruncDate('created_at')
        ).values('date').annotate(
            revenue=Sum('total_amount'),
            orders=Count('id')
        ).order_by('date')
        
        chart_labels = []
        chart_revenue = []
        chart_orders = []
        
        for entry in daily_sales:
            chart_labels.append(entry['date'].strftime('%b %d'))
            chart_revenue.append(float(entry['revenue'] or 0))
            chart_orders.append(entry['orders'])
        
        top_products = OrderItem.objects.filter(
            order__in=completed_orders
        ).values(
            'variant__product__name', 
            'variant__product__brand'
        ).annotate(
            total_qty=Sum('quantity'),
            total_revenue=Sum(F('quantity') * F('unit_price'))
        ).order_by('-total_revenue')[:10]
        
        top_by_qty = OrderItem.objects.filter(
            order__in=completed_orders
        ).values(
            'variant__product__name', 
            'variant__product__brand'
        ).annotate(
            total_qty=Sum('quantity'),
            total_revenue=Sum(F('quantity') * F('unit_price'))
        ).order_by('-total_qty')[:10]
        
        category_sales = OrderItem.objects.filter(
            order__in=completed_orders
        ).values(
            'variant__product__category'
        ).annotate(
            total_revenue=Sum(F('quantity') * F('unit_price'))
        ).order_by('-total_revenue')
        
        context.update({
            'total_revenue': total_revenue,
            'total_refunds': abs(total_refunds),
            'net_revenue': net_revenue,
            'total_cost': total_cost,
            'total_profit': total_profit,
            'profit_margin': profit_margin,
            'gross_profit': gross_profit,
            'gross_margin': gross_margin,
            'total_orders': total_orders,
            'total_items': total_items,
            'avg_order_value': avg_order_value,
            'chart_labels': json.dumps(chart_labels),
            'chart_revenue': json.dumps(chart_revenue),
            'chart_orders': json.dumps(chart_orders),
            'top_products': top_products,
            'top_by_qty': top_by_qty,
            'category_sales': category_sales,
            'start_date': start_date,
            'end_date': end_date,
        })
        
        return context
class SalesAssociateDashboardView(SalesAssociateRequiredMixin, TemplateView):
    template_name = 'core/sales_associate_dashboard.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        query = self.request.GET.get('q')
        if query:
            products = Product.objects.filter(
                Q(base_sku__icontains=query) |
                Q(name__icontains=query) |
                Q(variants__sku__icontains=query)
            ).distinct()
            context['products'] = products
            context['query'] = query
        return context


from .models import PromotionProduct
from .forms import PromotionForm

class PromotionListView(OwnerRequiredMixin, ListView):
    model = Promotion
    template_name = 'core/promotion_list.html'
    context_object_name = 'promotions'

    def get_queryset(self):
        qs = Promotion.objects.all()
        status_filter = self.request.GET.get('status')
        if status_filter == 'active':
            now = timezone.now()
            qs = qs.filter(is_active=True, start_date__lte=now, end_date__gte=now)
        elif status_filter == 'scheduled':
            now = timezone.now()
            qs = qs.filter(is_active=True, start_date__gt=now)
        elif status_filter == 'expired':
            now = timezone.now()
            qs = qs.filter(end_date__lt=now)
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['status_filter'] = self.request.GET.get('status', '')
        return context


class PromotionCreateView(OwnerRequiredMixin, CreateView):
    model = Promotion
    form_class = PromotionForm
    template_name = 'core/promotion_form.html'
    success_url = reverse_lazy('promotion_list')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        response = super().form_valid(form)
        
        products = form.cleaned_data.get('products')
        if products:
            for product in products:
                PromotionProduct.objects.create(promotion=self.object, product=product)
        
        messages.success(self.request, f"Promotion '{self.object.name}' created successfully!")
        return response


class PromotionUpdateView(OwnerRequiredMixin, UpdateView):
    model = Promotion
    form_class = PromotionForm
    template_name = 'core/promotion_form.html'
    success_url = reverse_lazy('promotion_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['editing'] = True
        return context

    def form_valid(self, form):
        response = super().form_valid(form)
        
        products = form.cleaned_data.get('products')
        self.object.product_links.all().delete()
        if products:
            for product in products:
                PromotionProduct.objects.create(promotion=self.object, product=product)
        
        messages.success(self.request, f"Promotion '{self.object.name}' updated successfully!")
        return response


class PromotionDeleteView(OwnerRequiredMixin, View):
    def post(self, request, pk):
        promotion = get_object_or_404(Promotion, pk=pk)
        name = promotion.name
        promotion.delete()
        messages.success(request, f"Promotion '{name}' deleted.")
        return redirect('promotion_list')


class PromotionToggleView(OwnerRequiredMixin, View):
    def post(self, request, pk):
        promotion = get_object_or_404(Promotion, pk=pk)
        promotion.is_active = not promotion.is_active
        promotion.save()
        status = "activated" if promotion.is_active else "deactivated"
        messages.success(request, f"Promotion '{promotion.name}' {status}.")
        return redirect('promotion_list')


from .models import User

class PersonnelListView(OwnerRequiredMixin, ListView):
    model = User
    template_name = 'core/personnel_list.html'
    context_object_name = 'employees'

    def get_queryset(self):
        return User.objects.exclude(is_superuser=True).order_by('role', 'username')


class PersonnelCreateView(OwnerRequiredMixin, View):
    template_name = 'core/personnel_form.html'

    def get(self, request):
        return render(request, self.template_name, {
            'roles': User.Role.choices
        })

    def post(self, request):
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '').strip()
        confirm_password = request.POST.get('confirm_password', '').strip()
        role = request.POST.get('role', User.Role.SALES_ASSOCIATE)
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()

        errors = []
        if not username:
            errors.append("Username is required.")
        elif User.objects.filter(username=username).exists():
            errors.append("Username already exists.")
        
        if not password:
            errors.append("Password is required.")
        elif len(password) < 4:
            errors.append("Password must be at least 4 characters.")
        
        if password != confirm_password:
            errors.append("Passwords do not match.")

        if errors:
            for error in errors:
                messages.error(request, error)
            return render(request, self.template_name, {
                'roles': User.Role.choices,
                'form_data': request.POST
            })

        user = User.objects.create_user(
            username=username,
            password=password,
            role=role,
            first_name=first_name,
            last_name=last_name
        )
        messages.success(request, f"Employee '{username}' created successfully as {user.get_role_display()}.")
        return redirect('personnel_list')


class PersonnelDeleteView(OwnerRequiredMixin, View):
    def post(self, request, pk):
        user = get_object_or_404(User, pk=pk)
        
        if user == request.user:
            messages.error(request, "You cannot delete your own account.")
            return redirect('personnel_list')
        
        if user.is_superuser:
            messages.error(request, "Cannot delete superuser accounts.")
            return redirect('personnel_list')
        
        username = user.username
        user.delete()
        messages.success(request, f"Employee '{username}' deleted.")
        return redirect('personnel_list')


from django.http import HttpResponse
from datetime import timedelta
from django.db.models import Count, Avg
from django.db.models.functions import TruncDate

class ReportExportView(OwnerRequiredMixin, View):
    def get(self, request):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            messages.error(request, "Excel export requires openpyxl. Install with: pip install openpyxl")
            return redirect('reports')
        
        days = int(request.GET.get('days', 30))
        end_date = timezone.now()
        start_date = end_date - timedelta(days=days)
        
        orders = Order.objects.filter(
            status=Order.Status.COMPLETED,
            created_at__range=(start_date, end_date)
        ).order_by('-created_at')
        
        wb = Workbook()
        
        ws = wb.active
        ws.title = "Summary"
        
        total_revenue = orders.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
        total_orders = orders.count()
        avg_order = total_revenue / total_orders if total_orders > 0 else Decimal('0')
        total_items = OrderItem.objects.filter(order__in=orders).aggregate(total=Sum('quantity'))['total'] or 0
        
        total_cost = Decimal('0')
        for order in orders:
            for item in order.items.all():
                if item.variant and item.variant.cost_price:
                    total_cost += item.variant.cost_price * item.quantity
        total_profit = total_revenue - total_cost
        
        header_font = Font(bold=True, size=14)
        header_fill = PatternFill(start_color="C0A464", end_color="C0A464", fill_type="solid")
        
        ws['A1'] = f"AURELION Sales Report ({days} Days)"
        ws['A1'].font = Font(bold=True, size=18)
        ws.merge_cells('A1:D1')
        
        ws['A3'] = "Metric"
        ws['B3'] = "Value"
        ws['A3'].font = header_font
        ws['B3'].font = header_font
        
        metrics = [
            ("Total Revenue", f"${total_revenue:,.2f}"),
            ("Total Cost", f"${total_cost:,.2f}"),
            ("Total Profit", f"${total_profit:,.2f}"),
            ("Profit Margin", f"{(total_profit/total_revenue*100):.1f}%" if total_revenue > 0 else "0%"),
            ("Total Orders", str(total_orders)),
            ("Total Items Sold", str(total_items)),
            ("Average Order Value", f"${avg_order:,.2f}"),
        ]
        
        for i, (metric, value) in enumerate(metrics, start=4):
            ws[f'A{i}'] = metric
            ws[f'B{i}'] = value
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        
        ws2 = wb.create_sheet("Orders")
        headers = ['Order Code', 'Date', 'Client', 'Items', 'Total', 'Cashier']
        for col, header in enumerate(headers, start=1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = header_font
        
        for row, order in enumerate(orders, start=2):
            ws2.cell(row=row, column=1, value=order.order_code)
            ws2.cell(row=row, column=2, value=order.created_at.strftime('%Y-%m-%d %H:%M'))
            ws2.cell(row=row, column=3, value=str(order.client) if order.client else 'Walk-in')
            ws2.cell(row=row, column=4, value=order.items.count())
            ws2.cell(row=row, column=5, value=float(order.total_amount))
            ws2.cell(row=row, column=6, value=order.created_by.username if order.created_by else '')
        
        ws3 = wb.create_sheet("Top Products")
        headers = ['Product', 'Brand', 'Qty Sold', 'Revenue']
        for col, header in enumerate(headers, start=1):
            cell = ws3.cell(row=1, column=col, value=header)
            cell.font = header_font
        
        top_products = OrderItem.objects.filter(order__in=orders).values(
            'variant__product__name', 'variant__product__brand'
        ).annotate(
            total_qty=Sum('quantity'),
            total_revenue=Sum(F('quantity') * F('unit_price'))
        ).order_by('-total_revenue')[:20]
        
        for row, item in enumerate(top_products, start=2):
            ws3.cell(row=row, column=1, value=item['variant__product__name'])
            ws3.cell(row=row, column=2, value=item['variant__product__brand'])
            ws3.cell(row=row, column=3, value=item['total_qty'])
            ws3.cell(row=row, column=4, value=float(item['total_revenue']))
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=aurelion_report_{days}days.xlsx'
        wb.save(response)
        return response


def health(request):
    """Health check endpoint that tests database connectivity"""
    from django.db import connection
    import os
    
    health_status = {
        "status": "ok",
        "database": "unknown",
        "env_vars": {}
    }
    
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT 1")
            cursor.fetchone()
        health_status["database"] = "connected"
    except Exception as e:
        health_status["database"] = f"error: {str(e)}"
        health_status["status"] = "error"
    
    return JsonResponse(health_status)
